"""PPE report routes."""
from __future__ import annotations
import os
import json
import io
from typing import Dict
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Request
from fastapi.templating import Jinja2Templates
from fastapi.responses import RedirectResponse
from modules.utils import require_roles
from fastapi.responses import StreamingResponse, JSONResponse

from modules.utils import send_email
from core.config import ANOMALY_ITEMS
from config import config

router = APIRouter()

BASE_DIR = Path(__file__).resolve().parent.parent

def init_context(config: dict, trackers: Dict[int, "PersonTracker"], redis_client, templates_path):
    global cfg, trackers_map, redis, templates
    cfg = config
    trackers_map = trackers
    redis = redis_client
    templates = Jinja2Templates(directory=templates_path)

@router.get('/ppe_report')
async def ppe_report_page(request: Request, status: str = ''):
    res = require_roles(request, ['admin'])
    if isinstance(res, RedirectResponse):
        return res
    statuses = []
    for item in cfg.get('track_ppe', []):
        statuses.append(item)
        if f'no_{item}' in ANOMALY_ITEMS:
            statuses.append(f'no_{item}')
    if cfg.get('track_misc', True):
        statuses.append('misc')
    return templates.TemplateResponse('ppe_report.html', {
        'request': request,
        'cfg': config,
        'status': status,
        'status_options': statuses,
    })

@router.get('/ppe_report_data')
async def ppe_report_data(start: str, end: str, status: str = '', min_conf: float | None = None, color: str | None = None):
    try:
        start_ts = int(datetime.fromisoformat(start).timestamp())
        end_ts = int(datetime.fromisoformat(end).timestamp())
    except Exception:
        return {"error": "invalid range"}
    entries = [json.loads(e) for e in redis.zrangebyscore('ppe_logs', start_ts, end_ts)]
    rows = []
    thresh = float(min_conf) if min_conf is not None else cfg.get('helmet_conf_thresh', 0.5)
    statuses = {s for s in status.split(',') if s}
    for e in entries:
        ts = e.get('ts')
        if statuses and e.get('status') not in statuses:
            continue
        if e.get('conf', 0) < thresh:
            continue
        if color and e.get('color') != color:
            continue
        path = e.get('path') or ''
        img_url = ''
        if path:
            fname = os.path.basename(path)
            img_url = f"/snapshots/{fname}"
        rows.append({
            'time': datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M'),
            'cam_id': e.get('cam_id', ''),
            'track_id': e.get('track_id', ''),
            'status': e.get('status', ''),
            'conf': float(e.get('conf', 0)),
            'color': e.get('color') or '',
            'image': img_url,
        })
    return {'rows': rows}

@router.get('/ppe_report/export')
async def ppe_report_export(start: str, end: str, status: str = '', min_conf: float | None = None, color: str | None = None):
    data = await ppe_report_data(start, end, status, min_conf, color)
    if 'error' in data:
        return JSONResponse(data, status_code=400)
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    wb = Workbook()
    ws = wb.active
    ws.append(['Time', 'Camera', 'Track', 'Status', 'Conf', 'Color', 'Image'])
    for row in data['rows']:
        ws.append([row['time'], row['cam_id'], row['track_id'], row['status'], round(row['conf'],2), row.get('color') or ''])
        img_path = row.get('image')
        if img_path:
            img_file = os.path.join(BASE_DIR, img_path.lstrip('/'))
            if os.path.exists(img_file):
                img = XLImage(img_file)
                img.width = 80
                img.height = 60
                ws.add_image(img, f'G{ws.max_row}')
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=ppe_report.xlsx'})

@router.post('/ppe_report/email')
async def ppe_report_email(start: str, end: str, status: str = '', min_conf: float | None = None, color: str | None = None, to: str | None = None):
    data = await ppe_report_data(start, end, status, min_conf, color)
    if 'error' in data:
        return JSONResponse(data, status_code=400)
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    wb = Workbook()
    ws = wb.active
    ws.append(['Time', 'Camera', 'Track', 'Status', 'Conf', 'Color', 'Image'])
    for row in data['rows']:
        ws.append([row['time'], row['cam_id'], row['track_id'], row['status'], round(row['conf'],2), row.get('color') or ''])
        img_path = row.get('image')
        if img_path:
            img_file = os.path.join(BASE_DIR, img_path.lstrip('/'))
            if os.path.exists(img_file):
                img = XLImage(img_file)
                img.width = 80
                img.height = 60
                ws.add_image(img, f'G{ws.max_row}')
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    recipients = [a.strip() for a in (to or cfg.get('email', {}).get('from_addr','')).split(',') if a.strip()]
    send_email('PPE Report', 'See attached report', recipients, None, cfg.get('email', {}), attachment=bio.getvalue(), attachment_name='ppe_report.xlsx', attachment_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return {'sent': True}
